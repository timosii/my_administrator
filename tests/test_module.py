import os
import asyncio

from aiogram import Bot
from aiogram.client.bot import DefaultBotProperties
from openpyxl.utils import get_column_letter
import pandas as pd
from openpyxl.styles import Alignment
from loguru import logger
from sqlalchemy import text
from openpyxl.drawing.image import Image
from app.database.database import session_maker
import os

# %%
import pandas as pd

# %%
pd.set_option('display.max_columns', None)

# %%
start_date = '2025-01-01'
end_date = '2025-02-10'

DATA_PATH = '/home/timosii/my_administrator/data/'

OUT_COLS = {
    'check_date': 'Дата проверки',
    'mfc_fio': 'ФИО проверяющего',
    'mfc_post': 'Должность',
    'mo_': 'МО',
    'fil_': 'Филиал',
    'comm_mfc': 'Комментарий МФЦ',
    'zone': 'Зона',
    'violation_name': 'Нарушение',
    'problem': 'Проблематика',
    'comm_mo': 'Комментарий МО',
    'mo_fio': 'ФИО сотрудника МО',
    'mo_post': 'Должность сотрудника МО',
    'violation_detected': 'Время обнаружения нарушения',
    'violation_fixed': 'Время исправления нарушения',
    'violation_pending': 'Время переноса нарушения',
    'pending_period': 'Срок переноса нарушения',
    'is_task': 'В рамках уведомления',
    'photo_id_mo': 'Фото МО',
}
async def test():
    async with session_maker() as session:
        query = """
        SELECT
            DATE(check_t.mfc_start AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow') AS check_date,
            CAST(check_t.check_id as varchar) as check_id,
            CAST(violation_found_t.violation_found_id as varchar) as violation_found_id,
            check_t.mfc_start AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS mfc_start,
            check_t.mfc_finish AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS mfc_finish,
            mfc_user.last_name || ' ' || mfc_user.first_name || ' ' || COALESCE(mfc_user.patronymic, '') AS MFC_FIO,
            mfc_user.post AS mfc_post,
            fil_t.mo_ AS mo_,
            fil_t.fil_ AS fil_,
            fil_t.fil_population AS fil_population,
            violation_found_t.comm_mfc AS comm_mfc,
            vio_dict_t.zone AS zone,
            vio_dict_t.violation_name AS violation_name,
            vio_dict_t.problem AS problem,
            violation_found_t.photo_id_mfc as photo_id_mfc,
            violation_found_t.photo_id_mo as photo_id_mo,
            violation_found_t.comm_mo AS comm_mo,
            mo_user.last_name || ' ' || mo_user.first_name || ' ' || COALESCE(mo_user.patronymic, '') AS MO_FIO,
            mo_user.post AS mo_post,
            violation_found_t.violation_detected AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS violation_detected,
            violation_found_t.violation_fixed AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS violation_fixed,
            check_t.is_task AS is_task,
            violation_found_t.is_pending AS is_pending,
            violation_found_t.violation_pending AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS violation_pending,
            DATE(violation_found_t.pending_period) AS pending_period
        FROM data.check as check_t
        left JOIN data.user AS mfc_user
            ON mfc_user.user_id = check_t.mfc_user_id
        JOIN data.violation_found as violation_found_t
            ON violation_found_t.check_id = check_t.check_id
        left JOIN dicts.filials as fil_t
            ON check_t.fil_ = fil_t.fil_
        left JOIN dicts.violations as vio_dict_t
            ON vio_dict_t.violation_dict_id = violation_found_t.violation_dict_id
        LEFT JOIN data.user AS mo_user
            ON mo_user.user_id = violation_found_t.mo_user_id
        WHERE DATE(check_t.mfc_start AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow') BETWEEN :start_date AND :end_date
        """
        result = await session.execute(text(query), {'start_date': start_date, 'end_date': end_date})
        rows = result.fetchall()
    if rows:
        df = pd.DataFrame(rows)
        df.columns = result.keys()
        df = df.sort_values(by='check_date')
        df_expanded = df['photo_id_mfc'].apply(pd.Series)
        df_expanded = df_expanded.rename(columns=lambda x: f'Фото МФЦ {x+1}')
        df = pd.concat([df, df_expanded], axis=1)
        df = df.drop(columns=['photo_id_mfc'])
        df['is_task'] = df['is_task'].map({False: 'Нет', True: 'Да'})
        df = df[list(OUT_COLS.keys()) + [col for col in df.columns if col.startswith('Фото МФЦ')]].rename(columns=OUT_COLS)
    else:
        logger.info('mfc report is empty')


    logger.info('get mfc report with photo')
    path_file = 'mfc_report_with_photo.xlsx'
    if os.path.exists(path_file):
        os.remove(path_file)
    # mfc_report_doc = FSInputFile(path_file)
    with pd.ExcelWriter(path_file, engine='openpyxl') as writer:
        sheet_name = 'Отчет по нарушениям'
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        worksheet = writer.sheets[sheet_name]

        for index, row in df.iterrows():
            max_height = 15
            for i, photo_col in enumerate(df_expanded.columns):
                photo_id = row[photo_col]
                if isinstance(photo_id, str):
                    try:
                        img_path = os.path.join(DATA_PATH, f'{photo_id}.jpg')
                        logger.debug(f'Попытка загрузить изображение: {img_path}')
                        if not os.path.exists(img_path):
                            logger.error(f'Файл не найден: {img_path}')
                        try:
                            img = Image(img_path)
                            logger.debug(f'Изображение успешно загружено: {img_path}')
                        except Exception as e:
                            logger.error(f'Ошибка при загрузке изображения {img_path}: {e}')
                            continue
                        img.width = 200
                        img.height = 150
                        cell = f'{get_column_letter(df.columns.get_loc(photo_col) + 1)}{index + 2}'
                        worksheet.add_image(img, cell)
                        worksheet[cell].value = None

                        column_letter = get_column_letter(df.columns.get_loc(photo_col) + 1)
                        worksheet.column_dimensions[column_letter].width = img.width / 7 + 2

                        if img.height > max_height:
                            max_height = img.height
                    except Exception as e:
                        logger.error(f'Ошибка при вставке изображения: {e}')

            photo_id_mo = row['Фото МО']
            if isinstance(photo_id_mo, str):
                try:
                    img_path = os.path.join(DATA_PATH, f'{photo_id_mo}.jpg')
                    img = Image(img_path)

                    img.width = 200
                    img.height = 150

                    cell = f'{get_column_letter(df.columns.get_loc("Фото МО") + 1)}{index + 2}'
                    worksheet.add_image(img, cell)
                    worksheet[cell].value = None
                    column_letter = get_column_letter(df.columns.get_loc('Фото МО') + 1)
                    worksheet.column_dimensions[column_letter].width = img.width / 7 + 2

                    if img.height > max_height:
                        max_height = img.height

                except Exception as e:
                    logger.error(f'Ошибка при вставке изображения: {e}')

            worksheet.row_dimensions[index + 2].height = max_height * 0.75

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)  # type: ignore

        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 10
        worksheet.column_dimensions['E'].width = 10
        worksheet.column_dimensions['F'].width = 35
        worksheet.column_dimensions['G'].width = 35
        worksheet.column_dimensions['H'].width = 35
        worksheet.column_dimensions['I'].width = 45
        worksheet.column_dimensions['J'].width = 25
        worksheet.column_dimensions['K'].width = 25
        worksheet.column_dimensions['L'].width = 25
        worksheet.column_dimensions['M'].width = 30
        worksheet.column_dimensions['N'].width = 30
        worksheet.column_dimensions['O'].width = 30
        worksheet.column_dimensions['P'].width = 25
        worksheet.column_dimensions['Q'].width = 25


if __name__=='__main__':
    asyncio.run(test())
