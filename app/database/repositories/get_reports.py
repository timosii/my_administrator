import asyncio
import os

import pandas as pd
from aiogram import Bot
from aiogram.client.bot import DefaultBotProperties
from aiogram.types import FSInputFile
from loguru import logger
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.config import settings
from app.database.database import session_maker


OUT_COLS_BASE = {
        'check_date': 'Дата проверки',
        'mfc_fio': 'ФИО проверяющего',
        'mfc_post': 'Должность',
        'mo_': 'МО',
        'fil_': 'Филиал',
        'comm_mfc': 'Комментарий МФЦ',
        'zone': 'Зона',
        'violation_name': 'Нарушение',
        'problem': 'Проблематика',
        'comm_mo': 'Комментарий МО',
        'mo_fio': 'ФИО сотрудника МО',
        'mo_post': 'Должность сотрудника МО',
        'violation_detected': 'Время обнаружения нарушения',
        'violation_fixed': 'Время исправления нарушения',
        'violation_pending': 'Время переноса нарушения',
        'pending_period': 'Срок переноса нарушения',
        'is_task': 'В рамках уведомления',
    }

problem_ids = [
    'AgACAgIAAxkBAAKlfGeQ12u5O7J_Teed_TxEd5UgPwMLAALC6DEbetWJSO9KHjaFCv86AQADAgADeQADNgQ',
    'AgACAgIAAxkBAAKloWeQ2OoVLV7lMJj1-IkDg0b9s5qKAALS6DEbetWJSAeXvkesKkhMAQADAgADeQADNgQ',
    'AgACAgIAAxkBAAKn8WeQ6Erypq8Xz_06wG1-ML9e8Pm_AALE-jEbm_uJSPdVbJcomxJHAQADAgADeQADNgQ',
    'AgACAgIAAxkBAAKoL2eQ6dS5aviQ1n9icDHb-v6bfVRgAALS6TEbetWJSNWjxDEqcQriAQADAgADeQADNgQ',
    'AgACAgIAAxkBAAKjXmeQxo9827AjRwWS9tzN-VV7KCUaAAJP7TEbeviJSKVTzpUW50FWAQADAgADeQADNgQ',
    'AgACAgIAAxkBAAKi1GeQxQOxma1immav3TLh98hHqxxFAAJF7TEbeviJSBohY6XWryRqAQADAgADeQADNgQ',
    'AgACAgIAAxkBAAKh6WeQvIbVXp-C2S_XC2yont_Eum4WAAIQ6DEb2iaISM6CXQRg-8jPAQADAgADeQADNgQ',
    'AgACAgIAAxkBAAKhVGeQtiwZ36seAryCTDpHCjtQQ_axAAKf7TEbLRGISPc6yFcH5kC1AQADAgADeQADNgQ',
]


async def get_mfc_report(start_date: str, end_date: str) -> FSInputFile:
    
    async with session_maker() as session:
        query = """
        SELECT
            DATE(check_t.mfc_start AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow') AS check_date,
            CAST(check_t.check_id as varchar) as check_id,
            CAST(violation_found_t.violation_found_id as varchar) as violation_found_id,
            check_t.mfc_start AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS mfc_start,
            check_t.mfc_finish AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS mfc_finish,
            mfc_user.last_name || ' ' || mfc_user.first_name || ' ' || COALESCE(mfc_user.patronymic, '') AS MFC_FIO,
            mfc_user.post AS mfc_post,
            fil_t.mo_ AS mo_,
            fil_t.fil_ AS fil_,
            fil_t.fil_population AS fil_population,
            violation_found_t.comm_mfc AS comm_mfc,
            vio_dict_t.zone AS zone,
            vio_dict_t.violation_name AS violation_name,
            vio_dict_t.problem AS problem,
            violation_found_t.photo_id_mfc as photo_id_mfc,
            violation_found_t.photo_id_mo as photo_id_mo,
            violation_found_t.comm_mo AS comm_mo,
            mo_user.last_name || ' ' || mo_user.first_name || ' ' || COALESCE(mo_user.patronymic, '') AS MO_FIO,
            mo_user.post AS mo_post,
            violation_found_t.violation_detected AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS violation_detected,
            violation_found_t.violation_fixed AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS violation_fixed,
            check_t.is_task AS is_task,
            violation_found_t.is_pending AS is_pending,
            violation_found_t.violation_pending AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS violation_pending,
            DATE(violation_found_t.pending_period) AS pending_period
        FROM data.check as check_t
        left JOIN data.user AS mfc_user
            ON mfc_user.user_id = check_t.mfc_user_id
        JOIN data.violation_found as violation_found_t
            ON violation_found_t.check_id = check_t.check_id
        left JOIN dicts.filials as fil_t
            ON check_t.fil_ = fil_t.fil_
        left JOIN dicts.violations as vio_dict_t
            ON vio_dict_t.violation_dict_id = violation_found_t.violation_dict_id
        LEFT JOIN data.user AS mo_user
            ON mo_user.user_id = violation_found_t.mo_user_id
        WHERE DATE(check_t.mfc_start AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow') BETWEEN :start_date AND :end_date
            AND photo_id_mfc IS NOT NULL
        """
        result = await session.execute(text(query), {'start_date': start_date, 'end_date': end_date})
        rows = result.fetchall()

    if rows:
        df = pd.DataFrame(rows)
        df.columns = result.keys()
        df = df.sort_values(by='check_date')
        df['is_task'] = df['is_task'].map({False: 'Нет', True: 'Да'})
        df = df[OUT_COLS_BASE.keys()].rename(columns=OUT_COLS_BASE)
    else:
        logger.info('mfc report is empty')
        return None

    logger.info('get mfc report')
    path_file = 'mfc_report.xlsx'
    if os.path.exists(path_file):
        os.remove(path_file)
    mfc_report_doc = FSInputFile(path_file)
    with pd.ExcelWriter(path_file) as writer:
        sheet_name = 'Отчет по нарушениям'
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        worksheet = writer.sheets[sheet_name]

        worksheet.set_column('A:A', 15)
        worksheet.set_column('B:B', 25)
        worksheet.set_column('C:C', 15)
        worksheet.set_column('D:D', 10)
        worksheet.set_column('E:E', 10)
        worksheet.set_column('F:F', 35)
        worksheet.set_column('G:G', 35)
        worksheet.set_column('H:H', 35)
        worksheet.set_column('I:I', 45)
        worksheet.set_column('J:J', 25)
        worksheet.set_column('K:K', 25)
        worksheet.set_column('L:L', 25)
        worksheet.set_column('M:M', 30)
        worksheet.set_column('N:N', 30)
        worksheet.set_column('O:O', 25)
        worksheet.set_column('P:P', 25)
        worksheet.set_column('Q:Q', 25)

    return mfc_report_doc


async def get_mfc_report_with_photo(start_date: str, end_date: str) -> FSInputFile:
    OUT_COLS_PHOTO = OUT_COLS_BASE.copy()
    OUT_COLS_PHOTO['photo_id_mo'] = 'Фото МО'

    async with session_maker() as session:
        query = """
        SELECT
            DATE(check_t.mfc_start AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow') AS check_date,
            CAST(check_t.check_id as varchar) as check_id,
            CAST(violation_found_t.violation_found_id as varchar) as violation_found_id,
            check_t.mfc_start AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS mfc_start,
            check_t.mfc_finish AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS mfc_finish,
            mfc_user.last_name || ' ' || mfc_user.first_name || ' ' || COALESCE(mfc_user.patronymic, '') AS MFC_FIO,
            mfc_user.post AS mfc_post,
            fil_t.mo_ AS mo_,
            fil_t.fil_ AS fil_,
            fil_t.fil_population AS fil_population,
            violation_found_t.comm_mfc AS comm_mfc,
            vio_dict_t.zone AS zone,
            vio_dict_t.violation_name AS violation_name,
            vio_dict_t.problem AS problem,
            violation_found_t.photo_id_mfc as photo_id_mfc,
            violation_found_t.photo_id_mo as photo_id_mo,
            violation_found_t.comm_mo AS comm_mo,
            mo_user.last_name || ' ' || mo_user.first_name || ' ' || COALESCE(mo_user.patronymic, '') AS MO_FIO,
            mo_user.post AS mo_post,
            violation_found_t.violation_detected AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS violation_detected,
            violation_found_t.violation_fixed AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS violation_fixed,
            check_t.is_task AS is_task,
            violation_found_t.is_pending AS is_pending,
            violation_found_t.violation_pending AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow' AS violation_pending,
            DATE(violation_found_t.pending_period) AS pending_period
        FROM data.check as check_t
        left JOIN data.user AS mfc_user
            ON mfc_user.user_id = check_t.mfc_user_id
        JOIN data.violation_found as violation_found_t
            ON violation_found_t.check_id = check_t.check_id
        left JOIN dicts.filials as fil_t
            ON check_t.fil_ = fil_t.fil_
        left JOIN dicts.violations as vio_dict_t
            ON vio_dict_t.violation_dict_id = violation_found_t.violation_dict_id
        LEFT JOIN data.user AS mo_user
            ON mo_user.user_id = violation_found_t.mo_user_id
        WHERE DATE(check_t.mfc_start AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Moscow') BETWEEN :start_date AND :end_date
            AND photo_id_mfc IS NOT NULL
        """
        result = await session.execute(text(query), {'start_date': start_date, 'end_date': end_date})
        rows = result.fetchall()
    if rows:
        df = pd.DataFrame(rows)
        df.columns = result.keys()
        df = df.sort_values(by='check_date')
        df_expanded = df['photo_id_mfc'].apply(pd.Series).reset_index(drop=True)
        df_expanded = df_expanded.rename(columns=lambda x: f'Фото МФЦ {x+1}')
        df_expanded = df_expanded.fillna('').astype('str')
        df = pd.concat([df, df_expanded], axis=1).reset_index(drop=True)
        df = df.drop(columns=['photo_id_mfc'])
        df['photo_id_mo'] = df['photo_id_mo'].fillna('').astype('str')
        df['is_task'] = df['is_task'].map({False: 'Нет', True: 'Да'})
        df = df[list(OUT_COLS_PHOTO.keys()) + [col for col in df.columns if col.startswith('Фото МФЦ')]].rename(columns=OUT_COLS_PHOTO)
    else:
        logger.info('mfc report is empty')
        return None

    logger.info('get mfc report with photo')
    path_file = 'mfc_report_with_photo.xlsx'
    if os.path.exists(path_file):
        os.remove(path_file)
    mfc_report_doc = FSInputFile(path_file)
    with pd.ExcelWriter(path_file, engine='openpyxl') as writer:
        sheet_name = 'Отчет по нарушениям'
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        worksheet = writer.sheets[sheet_name]
        count_photos = 0
        count_photos_downloaded = 0
        for index, row in df.iterrows():
            max_height = 15
            for photo_col in df_expanded.columns:
                photo_id = row[photo_col]
                if photo_id in problem_ids:
                    logger.debug('HERE IS PROBLEM_ID_MFC!!')
                if photo_id:
                    count_photos += 1
                    img_path = os.path.join(settings.DATA_PATH, f'{photo_id}.jpg')
                    logger.debug(f'Попытка загрузить изображение: {img_path}')
                    if not os.path.exists(img_path):
                        logger.error(f'Файл не найден: {img_path}')
                    try:
                        img = Image(img_path)
                        logger.debug(f'Изображение успешно загружено: {img_path}')
                    except Exception as e:
                        logger.error(f'Ошибка при загрузке изображения {img_path}: {e}')
                        continue
                    try:
                        img.width = 200
                        img.height = 150
                        cell = f'{get_column_letter(df.columns.get_loc(photo_col) + 1)}{index + 2}'
                        worksheet.add_image(img, cell)
                        worksheet[cell].value = None
                        logger.debug(f'Изображение успешно добавлено в ячейку: {img_path}')
                        count_photos_downloaded += 1
                    except Exception as e:
                        logger.error(f'Что-то не так с фото: {e}')
                        continue

                    column_letter = get_column_letter(df.columns.get_loc(photo_col) + 1)
                    worksheet.column_dimensions[column_letter].width = img.width / 7 + 2

                    if img.height > max_height:
                        max_height = img.height

            photo_id_mo = row['Фото МО']
            if photo_id_mo:
                if photo_id_mo in problem_ids:
                    logger.debug('HERE IS PROBLEM_ID_MO!!')
                count_photos += 1
                img_path = os.path.join(settings.DATA_PATH, f'{photo_id_mo}.jpg')
                logger.debug(f'Попытка загрузить изображение: {img_path}')
                if not os.path.exists(img_path):
                    logger.error(f'Файл не найден: {img_path}')
                try:
                    img = Image(img_path)
                    logger.debug(f'Изображение успешно загружено: {img_path}')
                except Exception as e:
                    logger.error(f'Ошибка при загрузке изображения {img_path}: {e}')
                    continue

                try:
                    img.width = 200
                    img.height = 150

                    cell = f'{get_column_letter(df.columns.get_loc("Фото МО") + 1)}{index + 2}'
                    worksheet.add_image(img, cell)
                    worksheet[cell].value = None
                    count_photos_downloaded += 1
                except Exception as e:
                    logger.error(f'Что-то не так с фото: {e}')
                    continue
                column_letter = get_column_letter(df.columns.get_loc('Фото МО') + 1)
                worksheet.column_dimensions[column_letter].width = img.width / 7 + 2

                if img.height > max_height:
                    max_height = img.height

            worksheet.row_dimensions[index + 2].height = max(max_height * 0.8, 25)

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)  # type: ignore

        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 10
        worksheet.column_dimensions['E'].width = 10
        worksheet.column_dimensions['F'].width = 35
        worksheet.column_dimensions['G'].width = 35
        worksheet.column_dimensions['H'].width = 35
        worksheet.column_dimensions['I'].width = 45
        worksheet.column_dimensions['J'].width = 25
        worksheet.column_dimensions['K'].width = 25
        worksheet.column_dimensions['L'].width = 25
        worksheet.column_dimensions['M'].width = 30
        worksheet.column_dimensions['N'].width = 30
        worksheet.column_dimensions['O'].width = 30
        worksheet.column_dimensions['P'].width = 25
        worksheet.column_dimensions['Q'].width = 25
        logger.debug(f'Всего фото: {count_photos}')
        logger.debug(f'Всего фото загружено: {count_photos_downloaded}')

    return mfc_report_doc


if __name__ == '__main__':
    doc = asyncio.run(get_mfc_report_with_photo(start_date='2025-01-19', end_date='2025-01-19'))
    bot = Bot(token=settings.BOT_TOKEN, default=DefaultBotProperties(parse_mode='HTML'))
    asyncio.run(bot.send_document(chat_id=settings.DEV_ID, document=doc))
